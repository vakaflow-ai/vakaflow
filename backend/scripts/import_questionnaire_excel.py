#!/usr/bin/env python3
"""
Import questions from Excel questionnaire files into submission_requirements table.
Maps questions to appropriate questionnaire types:
- TPRM- Questionnaire
- Vendor Security Questionnaire
- Sub Contractor Questionnaire
- Vendor Qualification
"""
import sys
import os
import re
import uuid
from datetime import datetime

# Add parent directory to path
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from app.core.database import SessionLocal
from app.models.submission_requirement import SubmissionRequirement
from app.models.tenant import Tenant
import logging

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

try:
    import openpyxl
except ImportError:
    logger.error("openpyxl not installed. Install with: pip install openpyxl")
    sys.exit(1)


def sanitize_field_name(catalog_id):
    """Convert catalog_id to valid snake_case field name
    Entity design: field_name is generated from catalog_id (short code), not from title/label
    """
    if not catalog_id:
        return None
    # Convert catalog_id (e.g., "REQ-COM-01") to field_name (e.g., "req_com_01")
    text = str(catalog_id).lower().strip()
    text = re.sub(r'[^a-z0-9_]+', '_', text)
    text = re.sub(r'_+', '_', text)  # Remove multiple underscores
    text = text.strip('_')  # Remove leading/trailing underscores
    # Ensure it starts with a letter
    if text and text[0].isdigit():
        text = 'req_' + text
    return text[:50] if text else None  # Limit to 50 chars for readability


def determine_questionnaire_type(filename, question_text, category=None):
    """Determine questionnaire type based on filename and question content"""
    filename_lower = filename.lower()
    question_lower = (question_text or '').lower()
    
    # Check filename first
    if 'tprm' in filename_lower or 'third-party' in filename_lower or 'third_party' in filename_lower:
        return "TPRM- Questionnaire"
    elif 'cybersecurity' in filename_lower or 'security' in filename_lower:
        return "Vendor Security Questionnaire"
    elif 'sub' in filename_lower and ('contractor' in filename_lower or 'contract' in filename_lower):
        return "Sub Contractor Questionnaire"
    elif 'ai' in filename_lower or 'qualification' in filename_lower:
        return "Vendor Qualification"
    
    # Check question content
    if any(keyword in question_lower for keyword in ['third-party', 'vendor risk', 'tprm', 'third party', 'supplier', 'vendor management']):
        return "TPRM- Questionnaire"
    elif any(keyword in question_lower for keyword in ['security', 'authentication', 'authorization', 'encryption', 'vulnerability', 'incident', 'cyber', 'firewall', 'malware']):
        return "Vendor Security Questionnaire"
    elif any(keyword in question_lower for keyword in ['sub-contractor', 'subcontractor', 'sub contractor', 'employee', 'support', 'maintenance', 'sla']):
        return "Sub Contractor Questionnaire"
    elif any(keyword in question_lower for keyword in ['compliance', 'certification', 'regulatory', 'gdpr', 'soc', 'iso', 'hipaa', 'pci', 'nist', 'qualification', 'certificate']):
        return "Vendor Qualification"
    
    # Default based on category
    if category:
        category_lower = category.lower()
        if 'security' in category_lower:
            return "Vendor Security Questionnaire"
        elif 'compliance' in category_lower:
            return "Vendor Qualification"
    
    return "Vendor Qualification"  # Default


def determine_category(question_text, questionnaire_type):
    """Determine category based on question content"""
    question_lower = (question_text or '').lower()
    
    if any(keyword in question_lower for keyword in ['security', 'authentication', 'authorization', 'encryption', 'vulnerability', 'incident', 'cyber', 'firewall', 'malware', 'access control']):
        return 'security'
    elif any(keyword in question_lower for keyword in ['compliance', 'certification', 'regulatory', 'gdpr', 'soc', 'iso', 'hipaa', 'pci', 'nist', 'audit']):
        return 'compliance'
    elif any(keyword in question_lower for keyword in ['technical', 'architecture', 'infrastructure', 'system', 'api', 'integration', 'deployment']):
        return 'technical'
    elif any(keyword in question_lower for keyword in ['business', 'pricing', 'licensing', 'support', 'sla', 'contract', 'vendor']):
        return 'business'
    
    # Default based on questionnaire type
    if questionnaire_type == "Vendor Security Questionnaire":
        return 'security'
    elif questionnaire_type == "Vendor Qualification":
        return 'compliance'
    elif questionnaire_type == "Sub Contractor Questionnaire":
        return 'business'
    else:
        return 'general'


def extract_questions_from_excel(filepath, questionnaire_type=None):
    """Extract questions from Excel file"""
    questions = []
    try:
        wb = openpyxl.load_workbook(filepath, data_only=True)
        
        # Try to find questions in all sheets
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            
            # Find header row (look for common question/question text headers)
            header_row = None
            question_col = None
            description_col = None
            category_col = None
            
            # Look for header row in first 15 rows
            for row_idx, row in enumerate(ws.iter_rows(max_row=15), start=1):
                for col_idx, cell in enumerate(row, start=1):
                    if not cell.value:
                        continue
                    cell_value = str(cell.value).lower().strip()
                    # Look for question column headers
                    if any(keyword in cell_value for keyword in ['question', 'q.', 'q ', 'requested information', 'assessment name']):
                        header_row = row_idx
                        question_col = col_idx
                    elif 'description' in cell_value or 'details' in cell_value:
                        description_col = col_idx
                    elif 'category' in cell_value or 'type' in cell_value:
                        category_col = col_idx
            
            # If no header found, try to detect by content
            # Look for rows that start with numbers or question patterns
            if not header_row:
                # Check first few rows to find where data starts
                for row_idx, row in enumerate(ws.iter_rows(max_row=20), start=1):
                    first_cell = row[0].value if row else None
                    if first_cell:
                        first_str = str(first_cell).strip()
                        # Check if it looks like a question number or question text
                        if (first_str and (
                            first_str[0].isdigit() or  # Starts with number
                            '?' in first_str or  # Contains question mark
                            any(word in first_str.lower() for word in ['describe', 'provide', 'explain', 'what', 'how', 'when', 'where', 'who', 'which', 'does', 'do', 'is', 'are', 'has', 'have', 'please'])
                        )):
                            header_row = row_idx - 1  # Header is row before
                            question_col = 1
                            break
                
                # Default: assume first column, skip first row
                if not header_row:
                    header_row = 1
                    question_col = 1
            
            # Extract questions from data rows
            # Skip empty rows at the start
            start_row = header_row + 1
            max_rows = ws.max_row
            
            for row_idx in range(start_row, min(start_row + 1000, max_rows + 1)):  # Limit to 1000 rows per sheet
                row = list(ws.iter_rows(min_row=row_idx, max_row=row_idx))[0] if row_idx <= max_rows else []
                if not row:
                    continue
                
                question_cell = row[question_col - 1] if question_col <= len(row) else None
                question_text = str(question_cell.value).strip() if question_cell and question_cell.value else None
                
                if not question_text or len(question_text) < 10:  # Skip very short or empty questions
                    continue
                
                # Skip non-question content (headers, instructions, metadata, etc.)
                question_lower = question_text.lower()
                skip_patterns = [
                    'field completed automatically',
                    'not supported',
                    'inventory -',
                    'attribute -',
                    'justification',
                    'enter a justification',
                    'modifications in this field',
                    'this question can only be updated',
                    'unique identifier',
                    'assessment name',
                    'assessment number',
                    'template name',
                    'organization name',
                    'tax id',
                    'sno.',
                    'response',
                    'drop files',
                    'enter response',
                    'type of question',
                    'multichoice -',
                    'looped attribute',
                    'assess_control',
                    'relationship and its',
                    'assess risk and its',
                    'pre-selected options',
                    'description',
                    'question description',
                ]
                if any(pattern in question_lower for pattern in skip_patterns):
                    continue
                
                # Skip if it's just a field type or metadata
                if question_text in ['Type of Question/Item', 'MultiChoice - multiple answer', 'MultiChoice - single answer', 
                                   'Looped Attribute (where parent question is multiselect)', 'Assess_Control and Its attributes',
                                   'Relationship and Its attributes', 'Assess Risk and Its attributes']:
                    continue
                
                # Skip if it's just metadata/header text
                if question_text.startswith(('AI', 'AIWF-', '1.', '2.', '3.', '4.', '5.')) and len(question_text) < 50:
                    # Might be a sub-question number, check if next cell has the actual question
                    if question_col < len(row):
                        next_cell = row[question_col] if question_col < len(row) else None
                        if next_cell and next_cell.value:
                            next_text = str(next_cell.value).strip()
                            if len(next_text) > 20:  # Use the next cell as the question
                                question_text = next_text
                                question_lower = question_text.lower()
                
                # Skip if it's just a category/type name without question mark or question words
                if len(question_text) < 20 and '?' not in question_text and not any(word in question_lower for word in ['describe', 'provide', 'explain', 'what', 'how', 'when', 'where', 'who', 'which', 'does', 'do', 'is', 'are', 'has', 'have', 'please', 'show', 'list']):
                    # Might be a category header, skip
                    continue
                
                # Clean up question text (remove leading numbers, colons, etc.)
                question_text = re.sub(r'^[A-Z0-9]+[:\-]\s*', '', question_text)  # Remove "AI01:", "1.", etc.
                question_text = question_text.strip()
                
                if len(question_text) < 10:
                    continue
                
                # Get description if available
                description = None
                if description_col and description_col <= len(row):
                    desc_cell = row[description_col - 1]
                    if desc_cell and desc_cell.value:
                        description = str(desc_cell.value).strip()
                
                # Get category if available
                category = None
                if category_col and category_col <= len(row):
                    cat_cell = row[category_col - 1]
                    if cat_cell and cat_cell.value:
                        category = str(cat_cell.value).strip()
                
                # Determine questionnaire type if not provided
                q_type = questionnaire_type or determine_questionnaire_type(
                    os.path.basename(filepath), 
                    question_text,
                    category
                )
                
                # Determine category
                cat = determine_category(question_text, q_type)
                
                questions.append({
                    'label': question_text[:255],  # Limit to 255 chars
                    'field_name': sanitize_field_name(question_text),
                    'description': description[:1000] if description else None,  # Limit description
                    'category': cat,
                    'questionnaire_type': q_type,
                    'field_type': 'textarea',  # Default to textarea for questionnaire questions
                    'is_required': False,
                    'allowed_response_types': ['text', 'file', 'url'],  # Questionnaire-style
                })
        
        logger.info(f"Extracted {len(questions)} questions from {os.path.basename(filepath)}")
        return questions
        
    except Exception as e:
        logger.error(f"Error reading {filepath}: {e}")
        return []


def import_questionnaires():
    """Import questions from all Excel files in questionaires folder"""
    db = SessionLocal()
    try:
        # Get default tenant (or first tenant)
        tenant = db.query(Tenant).first()
        if not tenant:
            logger.error("No tenant found. Please create a tenant first.")
            return
        
        logger.info(f"Using tenant: {tenant.name} ({tenant.id})")
        
        # Questionnaire folder path (at project root level)
        # Script is in backend/scripts/, so go up 2 levels to get to project root
        script_dir = os.path.dirname(os.path.abspath(__file__))
        backend_dir = os.path.dirname(script_dir)
        project_root = os.path.dirname(backend_dir)
        questionnaire_dir = os.path.join(project_root, 'questionaires')
        if not os.path.exists(questionnaire_dir):
            logger.error(f"Questionnaire directory not found: {questionnaire_dir}")
            return
        logger.info(f"Reading questionnaires from: {questionnaire_dir}")
        
        # Get existing field names to avoid duplicates
        existing_field_names = set(
            db.query(SubmissionRequirement.field_name)
            .filter(SubmissionRequirement.tenant_id == tenant.id)
            .all()
        )
        existing_field_names = {name[0] for name in existing_field_names if name[0]}
        
        logger.info(f"Found {len(existing_field_names)} existing requirements in database")
        
        # Process each Excel file
        excel_files = [
            ('TPRM_6925.xlsx', 'TPRM- Questionnaire'),
            ('Cybersecurity Vendor Questionnaire.xlsx', 'Vendor Security Questionnaire'),
            ('AI Questionnaire 2.0 - ETA 17 Sep.xlsx', 'Vendor Qualification'),
            ('8420 _ (Alert Enterprise, Inc.) Alert Enterprise - EARC Demand - 10.23.2025.xlsx', 'Vendor Security Questionnaire'),
        ]
        
        total_imported = 0
        total_skipped = 0
        
        for filename, default_type in excel_files:
            filepath = os.path.join(questionnaire_dir, filename)
            if not os.path.exists(filepath):
                logger.warning(f"File not found: {filepath}")
                continue
            
            logger.info(f"\nProcessing: {filename}")
            questions = extract_questions_from_excel(filepath, default_type)
            
            for q_data in questions:
                if not q_data['field_name']:
                    logger.warning(f"Skipping question with invalid field_name: {q_data['label'][:50]}")
                    total_skipped += 1
                    continue
                
                # Check for duplicates by field_name
                if q_data['field_name'] in existing_field_names:
                    logger.debug(f"Skipping duplicate: {q_data['field_name']}")
                    total_skipped += 1
                    continue
                
                # Check for duplicates by label (similar questions)
                existing_by_label = db.query(SubmissionRequirement).filter(
                    SubmissionRequirement.tenant_id == tenant.id,
                    SubmissionRequirement.label.ilike(f"%{q_data['label'][:50]}%")
                ).first()
                
                if existing_by_label:
                    logger.debug(f"Skipping similar question: {q_data['label'][:50]}")
                    total_skipped += 1
                    continue
                
                # Create new requirement
                requirement = SubmissionRequirement(
                    tenant_id=tenant.id,
                    label=q_data['label'],
                    field_name=q_data['field_name'],
                    field_type=q_data['field_type'],
                    description=q_data.get('description'),
                    category=q_data['category'],
                    questionnaire_type=q_data['questionnaire_type'],
                    is_required=q_data['is_required'],
                    allowed_response_types=q_data['allowed_response_types'],
                    source_type='manual',
                    source_name=f"Imported from {filename}",
                    is_auto_generated=False,
                    is_enabled=True,
                    is_active=True,
                )
                
                db.add(requirement)
                existing_field_names.add(q_data['field_name'])
                total_imported += 1
                logger.info(f"  ✓ Imported: {q_data['label'][:60]}... -> {q_data['questionnaire_type']}")
        
        db.commit()
        
        logger.info(f"\n{'='*60}")
        logger.info(f"Import Summary:")
        logger.info(f"  Total imported: {total_imported}")
        logger.info(f"  Total skipped (duplicates): {total_skipped}")
        
        # Show distribution by questionnaire type
        logger.info(f"\nDistribution by Questionnaire Type:")
        for q_type in ["TPRM- Questionnaire", "Vendor Security Questionnaire", "Sub Contractor Questionnaire", "Vendor Qualification"]:
            count = db.query(SubmissionRequirement).filter(
                SubmissionRequirement.tenant_id == tenant.id,
                SubmissionRequirement.questionnaire_type == q_type
            ).count()
            logger.info(f"  {q_type}: {count} requirements")
        
    except Exception as e:
        logger.error(f"Error importing questionnaires: {e}", exc_info=True)
        db.rollback()
        raise
    finally:
        db.close()


if __name__ == "__main__":
    import_questionnaires()
